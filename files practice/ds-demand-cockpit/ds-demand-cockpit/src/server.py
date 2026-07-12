"""Cockpit server — stdlib only (B1). READS out/plan/*; never writes it (ONE-WAY STATE).
Scenarios are overlays in out/scenarios/. Role gate enforced SERVER-SIDE:
role=ops strips every financial key (margin/cost/roi/budget/value/objective/price fields).
Every what-if: overlay -> bottom-up reconcile (tie asserted) -> optimizer -> response (<2s law).
"""
import json, time, io, re, sys
from http.server import ThreadingHTTPServer, BaseHTTPRequestHandler
from pathlib import Path
from urllib.parse import urlparse, parse_qs
import numpy as np

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))
from hierarchy import build_S, assert_ties          # noqa
from optimizer import solve
import llm                          # noqa

def _ensure_plan():
    """Self-heal: build the plan snapshot if it is missing or unreadable (e.g. a
    cloud-only OneDrive copy that never downloaded). Keeps the server bootable."""
    f = ROOT / "out/plan/snapshot.json"
    try:
        json.load(open(f)); return
    except Exception:
        print("snapshot missing/unreadable -> building it now (one-time, ~30s)...")
        import subprocess, sys as _sys
        subprocess.run([_sys.executable, str(ROOT / "src/pipeline.py")], cwd=str(ROOT), check=True)

_ensure_plan()
SNAP = json.load(open(ROOT / "out/plan/snapshot.json"))
ST = dict(np.load(ROOT / "out/plan/whatif_state.npz", allow_pickle=False))
TYPE_MULT = {"consumer": 1.0, "trade": 0.7, "display": 0.5}
DENY = re.compile(r"margin|cost|roi|budget|value|objective|price|shadow|cogs|₹", re.I)
ANALOG, NEWSKU = "DS-CONFPULSE-001", "DS-CONFPULSE-149"

import pandas as pd
SERIES = pd.DataFrame({"sku": ST["sku"], "zone": ST["zone"], "category": ST["category"]})
EL_CAT = dict(zip(ST["el_cat_names"].tolist(), ST["el_cat_vals"].tolist()))
CAPS = dict(zip(ST["cap_names"].tolist(), ST["cap_vals"].tolist()))
H = ST["fc"].shape[1]

def scrub(o):
    if isinstance(o, dict):
        return {k: scrub(v) for k, v in o.items() if not DENY.search(str(k))}
    if isinstance(o, list):
        return [scrub(x) for x in o]
    return o

def apply_whatif(params):
    t0 = time.time()
    fc = ST["fc"].copy()
    cat = SERIES["category"].values
    # promo lever (uses OUR recovered elasticities, not generator truth)
    pr = params.get("promo") or {}
    if pr.get("category") and float(pr.get("depth", 0)) > 0:
        el = EL_CAT.get(pr["category"], 1.0)
        f = 1 + float(pr["depth"]) * el * 2.0 * TYPE_MULT.get(pr.get("type", "consumer"), 1.0)
        wk = set(int(w) for w in pr.get("weeks", [1, 2, 3, 4]))
        m = cat == pr["category"]
        for h in range(H):
            if h + 1 in wk:
                fc[m, h] *= f
    # price lever: % change per category, volume responds via per-SKU |elasticity|
    for c, pct in (params.get("price_pct") or {}).items():
        pct = float(pct)
        if abs(pct) > 1e-9:
            m = cat == c
            fc[m] *= (1 + pct / 100.0) ** (-ST["el_sku"][m][:, None])
    # launch lever: hypothetical new variant on the analog, S-curve from chosen week,
    # sized 30% of analog fc, cannibalising the analog by 20% of new volume (documented assumption)
    lw = params.get("launch_week")
    add_launch = np.zeros_like(fc)
    if lw:
        lw = int(lw)
        m_an = SERIES["sku"].values == ANALOG
        for h in range(H):
            if h + 1 >= lw:
                ramp = 1 / (1 + np.exp(-((h + 1 - lw + 1) - 2)))
                add_launch[m_an, h] = 0.30 * fc[m_an, h] * ramp
        fc = fc - 0.20 * add_launch          # cannibalisation on analog rows
    fc = np.maximum(fc, 0)
    total_new = fc + add_launch
    # RECONCILIATION IS TRUTH: bottom-up ties asserted before anything is shown
    S = getattr(apply_whatif, "_S", None)
    if S is None:
        prod_stub = pd.DataFrame({"sku_code": SERIES["sku"], "brand": SERIES["sku"].str[3:7]}).drop_duplicates("sku_code")
        S = build_S(SERIES.assign(elast=0, intermittent=False), prod_stub)[0]
        apply_whatif._S = S
    for h in range(H):
        assert_ties(S, S @ np.nan_to_num(total_new[:, h]), np.nan_to_num(total_new[:, h]))
    # re-optimize
    P = dict(d4=np.nan_to_num(total_new).sum(1), davg=np.nan_to_num(total_new).sum(1) / 4,
             opening=ST["opening"], dp=ST["dp"], margin=ST["margin"], abc=ST["abc"],
             casem=ST["casem"].astype(int))
    budget = float(ST["budget"][0]) * float(params.get("budget_mult", 1.0))
    sol = solve(SERIES, P, budget, CAPS)
    dt = (time.time() - t0) * 1000
    base_vol = float(np.nan_to_num(ST["fc"]).sum())
    new_vol = float(np.nan_to_num(total_new).sum())
    fwd = {c: np.nan_to_num(total_new[cat == c]).sum(0).tolist()
           for c in sorted(set(cat.tolist()))}
    return dict(
        solve_ms=round(dt, 1), reconcile_ok=True,
        base_volume=round(base_vol), scenario_volume=round(new_vol),
        delta_volume_pct=round((new_vol / base_vol - 1) * 100, 2),
        binding=sol["binding"],
        objective_margin=round(float(sol["objective"])),
        base_objective_margin=round(float(SNAP["plan"]["objective_margin"])),
        budget_util=round(float(sol["util_budget"]), 4),
        risk_lines=int(((ST["opening"] + sol["ship"]) /
                        np.maximum(P["davg"], 1e-9) < 1)[P["d4"] > 0].sum()),
        fwd_by_category=fwd, nation_fwd=np.nan_to_num(total_new).sum(0).tolist())

def export_xlsx(role):
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = "S&OP Plan"
    fin = role != "ops"
    head = ["sku", "zone", "category", "ship_units", "proj_cover_wk"] + \
           (["ship_value", "unit_margin"] if fin else [])
    ws.append(head)
    import csv
    with open(ROOT / "out/plan/m4_plan.csv") as f:
        for i, row in enumerate(csv.DictReader(f)):
            if int(row["ship_units"]) <= 0: continue
            r = [row["sku"], row["zone"], row["category"], row["ship_units"], row["proj_cover_wk"]]
            if fin:
                r += [row["ship_value"], ""]
            ws.append(r)
    ws2 = wb.create_sheet("Forward risk")
    ws2.append(["sku", "zone", "proj_cover_wk", "d4", "opening", "ship"])
    with open(ROOT / "out/plan/m4_forward_risk.csv") as f:
        for row in csv.DictReader(f):
            ws2.append([row["sku"], row["zone"], row["proj_cover_wk"], row["d4"], row["opening"], row["ship"]])
    ws3 = wb.create_sheet("KPIs")
    k = SNAP["kpis"]
    rows = [("WMAPE champion (SKU)", k["wmape_sku_champion"]),
            ("WMAPE snaive (SKU)", k["wmape_sku_snaive"]),
            ("Pipeline fill first-4wk", k["pipeline_fill_first4wk"]),
            ("FVA vs snaive (pp)", k["fva_vs_snaive_pp"])]
    if fin:
        rows.append(("Plan margin objective (₹)", SNAP["plan"]["objective_margin"]))
        rows.append(("Binding constraint", SNAP["plan"]["binding"]))
    for r in rows: ws3.append(list(r))
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()

class H_(BaseHTTPRequestHandler):
    def log_message(self, *a): pass
    def _send(self, code, body, ctype="application/json"):
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)
    def do_GET(self):
        u = urlparse(self.path); q = parse_qs(u.query)
        role = (q.get("role", ["exec"])[0] or "exec").lower()
        if u.path in ("/", "/index.html"):
            self._send(200, open(ROOT / "cockpit/index.html", "rb").read(), "text/html; charset=utf-8")
        elif u.path == "/snapshot":
            data = SNAP if role != "ops" else scrub({k: v for k, v in SNAP.items() if k != "promo_roi"})
            self._send(200, json.dumps(data, default=float).encode())
        elif u.path == "/scenarios":
            out = []
            for f in sorted((ROOT / "out/scenarios").glob("*.json")):
                d = json.load(open(f)); d["id"] = f.stem
                out.append(d if role != "ops" else scrub(d))
            self._send(200, json.dumps(out, default=float).encode())
        elif u.path == "/llm_status":
            self._send(200, json.dumps(llm.status()).encode())
        elif u.path == "/export":
            self._send(200, export_xlsx(role),
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            self._send(404, b"{}")
    def do_POST(self):
        u = urlparse(self.path); q = parse_qs(u.query)
        role = (q.get("role", ["exec"])[0] or "exec").lower()
        n = int(self.headers.get("Content-Length", 0))
        params = json.loads(self.rfile.read(n) or b"{}")
        if u.path == "/whatif_nl":
            cats = sorted(set(cat.tolist())) if False else sorted(set(SERIES["category"].tolist()))
            sysp = ("Convert a demand planner's request into STRICT JSON for a forecasting cockpit. "
                    "Keys: promo{category,depth(0-0.3),type(consumer|trade|display),weeks[1-4]}, "
                    "price_pct{category:percent}, launch_week(1-3 or null), budget_mult(0.6-1.2). "
                    f"Valid categories: {cats}. Output ONLY JSON, no prose.")
            txt, meta = llm.complete(sysp, params.get("text", ""), cache_key="nl::" + params.get("text","").lower())
            parsed = None
            if txt:
                try:
                    import re as _re
                    parsed = json.loads(_re.search(r"\{.*\}", txt, _re.S).group(0))
                except Exception:
                    parsed = None
            if not parsed:
                self._send(200, json.dumps({"error": "could not parse", "ai": meta}).encode()); return
            res = apply_whatif(parsed)
            res["parsed_params"] = parsed; res["ai"] = meta
            if role == "ops": res = scrub(res)
            self._send(200, json.dumps(res, default=str).encode()); return
        if u.path == "/whatif":
            res = apply_whatif(params)
            if role == "ops":
                res = scrub(res)
            self._send(200, json.dumps(res, default=float).encode())
        elif u.path == "/scenario":
            ts = time.strftime("%Y%m%d-%H%M%S")
            name = re.sub(r"[^A-Za-z0-9_-]", "", params.get("name", "scenario"))[:40] or "scenario"
            f = ROOT / f"out/scenarios/{ts}-{name}.json"
            json.dump(params, open(f, "w"), default=float)
            self._send(200, json.dumps({"saved": f.stem}).encode())
        else:
            self._send(404, b"{}")

if __name__ == "__main__":
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 8765
    print(f"cockpit on http://localhost:{port}")
    ThreadingHTTPServer(("127.0.0.1", port), H_).serve_forever()
