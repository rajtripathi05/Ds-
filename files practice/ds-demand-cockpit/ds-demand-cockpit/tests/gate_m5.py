"""M5 gate: server up; what-if <2s (5 diverse scenarios); scenario save+list;
role gate = ops payloads contain ZERO financial keys anywhere; xlsx export both roles."""
import sys, json, time, subprocess, re, io
from urllib.request import urlopen, Request

PORT = 8791
srv = subprocess.Popen([sys.executable, "src/server.py", str(PORT)],
                       stdout=subprocess.DEVNULL, stderr=subprocess.STDOUT)
def get(path):
    return urlopen(f"http://127.0.0.1:{PORT}{path}", timeout=30).read()
def post(path, obj):
    return urlopen(Request(f"http://127.0.0.1:{PORT}{path}", data=json.dumps(obj).encode(),
                           method="POST"), timeout=30).read()
try:
    for _ in range(60):
        try: get("/snapshot?role=exec"); break
        except Exception: time.sleep(0.5)
    checks = []
    html = get("/").decode()
    checks.append(("cockpit html serves", "ds-demand-cockpit" in html))
    ex = json.loads(get("/snapshot?role=exec"))
    checks.append(("exec sees financial layer", "objective_margin" in json.dumps(ex)))
    ops = json.loads(get("/snapshot?role=ops"))
    DENY = re.compile(r"margin|cost|roi|budget|value|objective|price|shadow|cogs", re.I)
    def leak(o, path=""):
        if isinstance(o, dict):
            for k, v in o.items():
                if DENY.search(str(k)): return path + "/" + str(k)
                r = leak(v, path + "/" + str(k))
                if r: return r
        if isinstance(o, list):
            for i, v in enumerate(o):
                r = leak(v, f"{path}[{i}]")
                if r: return r
        return None
    lk = leak(ops)
    checks.append((f"ops payload zero financial keys (leak={lk})", lk is None))
    # what-if timing: 5 diverse scenarios
    scens = [
        {"promo": {"category": "Confectionery", "depth": .25, "type": "consumer", "weeks": [1,2,3,4]}},
        {"price_pct": {"Beverages": 8}},
        {"promo": {"category": "Spices", "depth": .15, "type": "trade", "weeks": [2,3]},
         "price_pct": {"Dairy": -5}},
        {"launch_week": 1, "budget_mult": 0.8},
        {"promo": {"category": "MouthFreshener", "depth": .3, "type": "display", "weeks": [1]},
         "price_pct": {"Confectionery": 5}, "launch_week": 2, "budget_mult": 0.7},
    ]
    times = []
    for s in scens:
        t0 = time.time()
        r = json.loads(post("/whatif?role=exec", s))
        wall = (time.time() - t0) * 1000
        times.append(wall)
        assert r["reconcile_ok"] and "binding" in r
    checks.append((f"5 what-ifs reconcile+solve, max wall {max(times):.0f} ms (<2000)",
                   max(times) < 2000))
    # scenario overlay honesty: promo raises Confectionery volume
    r = json.loads(post("/whatif?role=exec", scens[0]))
    checks.append(("promo scenario raises volume", r["delta_volume_pct"] > 1))
    r2 = json.loads(post("/whatif?role=exec", {"price_pct": {"Beverages": 8}}))
    checks.append(("price rise lowers volume", r2["delta_volume_pct"] < 0))
    # ops whatif has no margin/budget keys
    r3 = json.loads(post("/whatif?role=ops", scens[0]))
    checks.append(("ops what-if scrubbed", leak(r3) is None))
    # scenarios save + list
    post("/scenario", {"name": "gate-A", "params": scens[0], "result": {"delta_volume_pct": r["delta_volume_pct"], "binding": r["binding"], "solve_ms": r["solve_ms"]}})
    post("/scenario", {"name": "gate-B", "params": scens[3], "result": {"delta_volume_pct": 0, "binding": "x", "solve_ms": 1}})
    lst = json.loads(get("/scenarios?role=exec"))
    checks.append(("scenario save+list >=2", len(lst) >= 2))
    # export xlsx both roles
    xe = get("/export?role=exec"); xo = get("/export?role=ops")
    checks.append(("xlsx exec magic+size", xe[:2] == b"PK" and len(xe) > 4000))
    checks.append(("xlsx ops magic+size", xo[:2] == b"PK" and len(xo) > 3000))
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xo)); heads = [c.value for c in wb["S&OP Plan"][1]]
    checks.append(("ops xlsx has no financial columns",
                   not any(h and re.search(r"margin|value|cost", str(h), re.I) for h in heads)))
    wb2 = load_workbook(io.BytesIO(xe)); heads2 = [c.value for c in wb2["S&OP Plan"][1]]
    checks.append(("exec xlsx has financial columns",
                   any(h and "value" in str(h) for h in heads2)))
    ok = all(c[1] for c in checks)
    for name, passed in checks:
        print(("✅" if passed else "❌"), name)
    json.dump({"checks": [(n, bool(p)) for n, p in checks],
               "whatif_wall_ms": [round(t) for t in times]},
              open("out/plan/m5_gate.json", "w"), indent=1)
    print("M5 GATE:", "GREEN" if ok else "RED")
    sys.exit(0 if ok else 1)
finally:
    srv.terminate()
